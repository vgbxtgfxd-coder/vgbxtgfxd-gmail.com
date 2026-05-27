"""
Парсер Excel-файлов с отчётами 1С ОСВ по номенклатуре и заказам.

Структура файла (на основе скриншота):
- Строка 2: Параметры отчёта (период, отбор)
- Строка 5: Заголовок "Номенклатура труб" | в ед. изм | в кг | в м
- Строка 6-7: Подзаголовки с колонками
- Строка 9+: Данные

Колонки:
A (1) - Номенклатура труб (иерархия: группа → номенклатура → склад → заказ)
B (2) - Номенклатура
C (3) - Склад
D (4) - Счёт
E (5) - Начальный остаток (ед)
F (6) - Оборот дт (ед)
G (7) - Оборот кт (ед)
H (8) - Конечный остаток (ед)
I (9) - Начальный остаток (кг)
J (10) - Оборот дт (кг)
K (11) - Оборот кт (кг)
L (12) - Конечный остаток (кг)
M (13) - Начальный остаток (м)
N (14) - Оборот дт (м)
O (15) - Оборот кт (м)
P (16) - Конечный остаток (м)
"""
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def open_workbook(file_path: str):
    """
    Открывает Excel-файл (.xlsx или .xls) и возвращает список строк.
    Каждая строка — список значений ячеек.
    """
    path = Path(file_path)
    suffix = path.suffix.lower()

    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        wb.close()
        return rows

    elif suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        rows = []
        for row_idx in range(ws.nrows):
            row_values = []
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                # xlrd возвращает пустые строки для пустых ячеек
                if val == "":
                    val = None
                row_values.append(val)
            rows.append(row_values)
        return rows

    else:
        raise ValueError(f"Неподдерживаемый формат: {suffix}")


def parse_stock_report(file_path: str) -> list[dict]:
    """
    Парсит Excel-файл отчёта ОСВ по номенклатуре и заказам.
    Поддерживает .xls и .xlsx.

    Returns:
        Список словарей с данными по каждой номенклатурной позиции (группы верхнего уровня).
    """
    path = Path(file_path)
    if not path.exists():
        logger.error(f"Файл не найден: {file_path}")
        return []

    try:
        rows = open_workbook(file_path)

        stock_items = []
        current_group = None

        # Определяем начало данных (ищем строку после заголовков)
        data_start_row = find_data_start(rows)
        if data_start_row is None:
            logger.warning("Не удалось определить начало данных")
            data_start_row = 9  # значение по умолчанию из скриншота

        logger.info(f"Начало данных: строка {data_start_row}")

        for row in rows[data_start_row - 1:]:  # rows 0-indexed, data_start_row 1-indexed
            # Получаем значения ячеек (до 16 колонок)
            cells = row[:16] if len(row) >= 16 else row + [None] * (16 - len(row))

            # Пропускаем полностью пустые строки
            if all(c is None for c in cells):
                continue

            name_col = cells[0]  # Колонка A — Номенклатура труб

            if name_col is None:
                continue

            name = str(name_col).strip()
            if not name:
                continue

            # Определяем уровень иерархии по отступу/форматированию
            # Группы обычно выделены жирным и имеют итоговые значения
            # Конкретные позиции имеют детализацию по складам

            # Проверяем есть ли числовые данные в строке
            end_qty = safe_float(cells[7])    # H - Конечный остаток (ед)
            end_kg = safe_float(cells[11])    # L - Конечный остаток (кг)
            end_m = safe_float(cells[15])     # P - Конечный остаток (м)

            start_qty = safe_float(cells[4])  # E - Начальный остаток (ед)
            turn_dt = safe_float(cells[5])    # F - Оборот дт
            turn_kt = safe_float(cells[6])    # G - Оборот кт

            start_kg = safe_float(cells[8])   # I - Начальный остаток (кг)
            start_m = safe_float(cells[12])   # M - Начальный остаток (м)

            # Определяем: это группа (итоговая строка номенклатуры) или детализация
            # Группы имеют данные в колонке A и итоговые числа
            # Детали — это строки со складами и заказами под группой

            warehouse_col = cells[2]  # C - Склад
            account_col = cells[3]    # D - Счёт

            # Если есть имя в A, нет склада в C, и есть конечный остаток —
            # это итоговая строка номенклатуры (группа)
            is_group_total = (
                name
                and warehouse_col is None
                and account_col is None
                and (end_qty or end_kg or end_m)
            )

            # Если есть склад — это детализация
            is_detail = warehouse_col is not None

            if is_group_total:
                item = {
                    "name": name,
                    "start_qty": format_number(start_qty),
                    "turn_dt_qty": format_number(turn_dt),
                    "turn_kt_qty": format_number(turn_kt),
                    "end_qty": format_number(end_qty),
                    "start_kg": format_number(start_kg),
                    "end_kg": format_number(end_kg),
                    "start_m": format_number(start_m),
                    "end_m": format_number(end_m),
                }
                stock_items.append(item)
                current_group = name

        logger.info(f"Распарсено позиций: {len(stock_items)}")
        return stock_items

    except Exception as e:
        logger.exception(f"Ошибка парсинга Excel: {e}")
        return []


def find_data_start(rows: list) -> int | None:
    """Ищет строку начала данных (после заголовков). Возвращает 1-indexed номер строки."""
    for row_num, row in enumerate(rows[:20], start=1):
        for cell in row:
            if cell and "Заказ клиента" in str(cell):
                return row_num + 1

    for row_num, row in enumerate(rows[:20], start=1):
        for cell in row:
            if cell and "Номенклатура труб" in str(cell):
                return row_num + 4

    return None


def safe_float(value) -> float | None:
    """Безопасное преобразование в float"""
    if value is None:
        return None
    try:
        result = float(value)
        return result if result != 0 else None
    except (ValueError, TypeError):
        return None


def format_number(value: float | None) -> str:
    """Форматирует число для отображения"""
    if value is None:
        return ""
    if value == int(value):
        return str(int(value))
    return f"{value:.2f}"


def parse_stock_report_detailed(file_path: str) -> list[dict]:
    """
    Расширенный парсинг — включает детализацию по складам и заказам.
    """
    path = Path(file_path)
    if not path.exists():
        return []

    try:
        rows = open_workbook(file_path)
        items = []
        current_group = None
        data_start_row = find_data_start(rows) or 9

        for row in rows[data_start_row - 1:]:
            cells = row[:16] if len(row) >= 16 else row + [None] * (16 - len(row))

            if all(c is None for c in cells):
                continue

            name = str(cells[0]).strip() if cells[0] else ""
            warehouse = str(cells[2]).strip() if cells[2] else ""
            account = str(cells[3]).strip() if cells[3] else ""

            end_qty = safe_float(cells[7])
            end_kg = safe_float(cells[11])
            end_m = safe_float(cells[15])

            if not name and not warehouse:
                continue

            item = {
                "group": current_group,
                "name": name if name else current_group,
                "warehouse": warehouse,
                "account": account,
                "end_qty": format_number(end_qty),
                "end_kg": format_number(end_kg),
                "end_m": format_number(end_m),
                "is_group": bool(name and not warehouse and not account),
            }

            if name and not warehouse and not account and (end_qty or end_kg or end_m):
                current_group = name

            items.append(item)

        return items

    except Exception as e:
        logger.exception(f"Ошибка парсинга: {e}")
        return []


if __name__ == "__main__":
    # Тест парсинга
    import sys

    if len(sys.argv) > 1:
        result = parse_stock_report(sys.argv[1])
        print(f"\nНайдено позиций: {len(result)}\n")
        for item in result[:10]:
            print(f"  {item['name']}: {item['end_qty']} ед | {item['end_kg']} кг | {item['end_m']} м")
    else:
        print("Использование: python parser_excel.py <path_to_excel>")
